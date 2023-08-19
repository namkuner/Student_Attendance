#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import os
sys.path.append(os.path.join(sys.path[0], 'MTCNN'))
import argparse
import torch
from torchvision import transforms as trans
from PIL import Image, ImageDraw, ImageFont
import numpy as np
from util import *
from align_trans import *
from MTCNN import create_mtcnn_net
from face_model import MobileFaceNet, l2_norm
from facebank import load_facebank, prepare_facebank
import cv2
import time
from anti_spoofing import predict
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill, colors
import os
import datetime
from openpyxl.utils import get_column_letter
day = datetime.date.today()
import torch.nn.functional as F
def cos_distance(vector1,vectors_set):
    return 1 - F.cosine_similarity(vector1, vectors_set)
def calculate_cosine_distance(source_representation, test_representation):
    source_representation = torch.tensor(source_representation)  # shape: (2, 512)
    test_representation = torch.tensor(test_representation)  # shape: (40, 512)

    # Tính cosine similarity
    cosine_similarity = torch.cosine_similarity(source_representation.unsqueeze(1), test_representation.unsqueeze(0), dim=-1)

    # Tính cosine distance
    cosine_distance = 1 - cosine_similarity

    return cosine_distance
def resize_image(img, scale):
    """
        resize image
    """
    height, width, channel = img.shape
    new_height = int(height * scale)     # resized new height
    new_width = int(width * scale)       # resized new width
    new_dim = (new_width, new_height)
    img_resized = cv2.resize(img, new_dim, interpolation=cv2.INTER_LINEAR)      # resized image
    return img_resized

def attendance(lop="D21CQCN01"):
    #lấy tên thiết bị
    diem_danh =[]
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    #xây dựng mô hình
    detect_model = MobileFaceNet(512).to(device)  # embeding size is 512 (feature vector)
    #lấy weight
    detect_model.load_state_dict(torch.load('Weights/MobileFace_Net', map_location=lambda storage, loc: storage))
    print('MobileFaceNet face detection model generated')
    #chuyển sang chế độ test. ở chế độ này thì model sẽ không tính đạo hàm bỏ qua batchnorm và dropout
    detect_model.eval()
    #lấy vector nhúng và tên của các nhãn
    targets, names,mssv = load_facebank(path='D21CQCN01')
    print('facebank loaded')
    #set camera
    cap = cv2.VideoCapture(0)
    # cap.set(cv2.CAP_PROP_AUTO_EXPOSURE, 1)
    # cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
    # cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
    # paramater
    scale = 0.5
    mini_face =30
    tta =True
    threshold= 60
    score_label =True
    while True:
        isSuccess, frame = cap.read()
        if isSuccess:
            try:
                start_time = time.time()
                input = resize_image(frame, scale)
                bboxes, landmarks = create_mtcnn_net(input, mini_face, device, p_model_path='Weights/pnet_Weights',
                                                     r_model_path='Weights/rnet_Weights',
                                                     o_model_path='Weights/onet_Weights')
                if bboxes != []:
                    bboxes = bboxes / scale
                    landmarks = landmarks / scale

                faces = Face_alignment(frame, default_square=True, landmarks=landmarks)

                embs = []

                test_transform = trans.Compose([
                                trans.ToTensor(),
                                trans.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5])])

                for img in faces:

                    if tta:
                        mirror = cv2.flip(img,1)
                        emb = detect_model(test_transform(img).to(device).unsqueeze(0))
                        emb_mirror = detect_model(test_transform(mirror).to(device).unsqueeze(0))
                        embs.append(l2_norm(emb + emb_mirror))
                    else:
                        embs.append(detect_model(test_transform(img).to(device).unsqueeze(0)))
                source_embs = torch.cat(embs)  # number of detected faces x 512

                print(targets.shape)
                diff = source_embs.unsqueeze(-1) - targets.transpose(1, 0).unsqueeze(0) # i.e. 3 x 512 x 1 - 1 x 512 x 2 = 3 x 512 x 2
                dist = torch.sum(torch.pow(diff, 2), dim=1) # number of detected faces x numer of target faces
                print("dist",dist)
                cos = cos_distance(source_embs,targets)
                # print("cos",cos)
                # minimum, min_idx = torch.min(dist, dim=1) # min and idx for each row
                print("min_idx",min_idx)
                min_idx[minimum > ((threshold-156)/(-80))] = -1  # if no match, set idx to -1
                # cos_dis = calculate_cosine_distance(source_embs,targets)
                # print(cos_dis)
                # minimum, min_idx = torch.min(cos_dis, dim=1) # min and idx for each row
                # print("min_idx",min_idx)
                # min_idx[minimum > ((threshold-156)/(-80))] = -1  # if no match, set idx to -1
                score = minimum
                results = min_idx
                diem_danh.append(mssv[results])
                print(score/torch.mean(dist, dim=1))
                # convert distance to score dis(0.7,1.2) to score(100,60)
                score_100 = torch.clamp(score*-80+156,0,100)

                image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                draw = ImageDraw.Draw(image)
                font = ImageFont.truetype('Weights/simkai.ttf', 30)

                FPS = 1.0 / (time.time() - start_time)
                draw.text((10, 10), 'FPS: {:.1f}'.format(FPS), fill=(0, 0, 0), font=font)

                for i, b in enumerate(bboxes):
                    draw.rectangle([(b[0], b[1]), (b[2], b[3])], outline='blue', width=5)
                    if score_label:

                        draw.text((int(b[0]), int(b[1]-25)), names[results[i] ] + ' score:{:.0f}'.format(score_100[i]), fill=(255,255,0), font=font)
                    else:
                        draw.text((int(b[0]), int(b[1]-25)), names[results[i] ], fill=(255,255,0), font=font)
                        print(names[results[i]])

                for p in landmarks:
                    for i in range(5):
                        draw.ellipse([(p[i] - 2.0, p[i + 5] - 2.0), (p[i] + 2.0, p[i + 5] + 2.0)], outline='blue')

                frame = cv2.cvtColor(np.asarray(image), cv2.COLOR_RGB2BGR)

            except:
                print('detect error')

            cv2.imshow('video', frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break
    diem_danh =set(diem_danh)
    print(diem_danh)
    print(names)
    cap.release()
    cv2.destroyAllWindows()
def load_model(name_model):
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    #xây dựng mô hình
    detect_model = MobileFaceNet(512).to(device)  # embeding size is 512 (feature vector)
    #lấy weight
    detect_model.load_state_dict(torch.load('Weights/MobileFace_Net', map_location=lambda storage, loc: storage))
    print('MobileFaceNet face detection model generated')
    detect_model.eval()
    return detect_model
def inference(frame,targets, names, mssv,detect_model,pnet,rnet,onet,device):
    scale = 0.5
    mini_face =50
    tta =True
    threshold= 60
    score_label =True
    res=[]
    start_time = time.time()
    input = resize_image(frame, scale)
    bboxes, landmarks = create_mtcnn_net(input, mini_face, device, pnet,
                                        rnet,
                                         onet)
    if bboxes.size != 0:
        bboxes = bboxes / scale
        landmarks = landmarks / scale
    faces = Face_alignment(frame, default_square=True, landmarks=landmarks)
    embs = []

    test_transform = trans.Compose([
        trans.ToTensor(),
        trans.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5])])

    for img in faces:

        if tta:
            mirror = cv2.flip(img, 1)
            emb = detect_model(test_transform(img).to(device).unsqueeze(0))
            emb_mirror = detect_model(test_transform(mirror).to(device).unsqueeze(0))
            embs.append(l2_norm(emb + emb_mirror))
        else:
            embs.append(detect_model(test_transform(img).to(device).unsqueeze(0)))
    source_embs = torch.cat(embs)  # number of detected faces x 512

    diff = source_embs.unsqueeze(-1) - targets.transpose(1, 0).unsqueeze(0)  # i.e. 3 x 512 x 1 - 1 x 512 x 2 = 3 x 512 x 2
    dist = torch.sum(torch.pow(diff, 2), dim=1)  # number of detected faces x numer of target faces
    print("educlid",dist)
    # print("dist", dist)
    minimum, min_idx = torch.min(dist, dim=1)  # min and idx for each row
    print("minimum",minimum)
    min_idx[minimum > ((threshold - 156) / (-80))] = -1  # if no match, set idx to -1
    # cos_dis = calculate_cosine_distance(source_embs,targets)
    # print(cos_dis)
    # minimum, min_idx = torch.min(cos_dis, dim=1) # min and idx for each row
    # print("min_idx",min_idx)
    # min_idx[minimum > ((threshold-156)/(-80))] = -1  # if no match, set idx to -1
    # similarity_scores = F.cosine_similarity(vector1, vectors_set)

    score = minimum
    results = min_idx
    for id in min_idx:
        if id != -1:
            res.append(mssv[id])
    # print(score / torch.mean(dist, dim=1))
    # convert distance to score dis(0.7,1.2) to score(100,60)
    score_100 = torch.clamp(score * -80 + 156, 0, 100)
    image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype('Weights/times.ttf', 30)

    FPS = 1.0 / (time.time() - start_time)
    draw.text((10, 10), 'FPS: {:.1f}'.format(FPS), fill=(0, 0, 0), font=font)

    for i, b in enumerate(bboxes):
        draw.rectangle([(b[0], b[1]), (b[2], b[3])], outline='blue', width=5)
        if score_label:
            if score_100[i]>70:
                draw.text((int(b[0]), int(b[1] - 25)), names[results[i]] + ' score:{:.0f}'.format(score_100[i]),
                      fill=(255, 255, 0), font=font)
            else:
                # results[i]="unknow"
                draw.text((int(b[0]), int(b[1] - 25)), "unknow", fill=(255, 255, 0), font=font)
        else:
            draw.text((int(b[0]), int(b[1] - 25)), names[results[i]], fill=(255, 255, 0), font=font)
            print(names[results[i]])

    for p in landmarks:
        for i in range(5):
            draw.ellipse([(p[i] - 2.0, p[i + 5] - 2.0), (p[i] + 2.0, p[i + 5] + 2.0)], outline='blue')
    frame = cv2.cvtColor(np.asarray(image), cv2.COLOR_RGB2BGR)
    return frame, res,score_100
def detec_with_face_spoofing(frame,targets, names, mssv,detect_model,pnet,rnet,onet,device,face_spoofing):
    scale = 0.5
    mini_face =30
    tta =True
    threshold= 60
    score_label =True
    res=[]
    start_time = time.time()
    input = resize_image(frame, scale)
    pre_box, pre_lank = create_mtcnn_net(input, mini_face, device, pnet,
                                        rnet,
                                         onet)


    bboxes =[]
    landmarks=[]
    fake_face = []


    for a in range(len(pre_box)):
        bbox = pre_box[a][:-1]
        prediction = predict(input,bbox,face_spoofing)
        if np.argmax(prediction)==1 or np.argmax(prediction)==0:
            bboxes.append(pre_box[a])
            landmarks.append(pre_lank[a])
        elif np.argmax(prediction)==2 :
            fake_face.append(pre_box[a])
    bboxes = np.array(bboxes)
    landmarks =np.array(landmarks)
    fake_face =np.array(fake_face)
    if fake_face.size!=0:
        fake_face= fake_face/scale
    if bboxes.size != 0:
        bboxes = bboxes / scale
        landmarks = landmarks / scale
    faces = Face_alignment(frame, default_square=True, landmarks=landmarks)
    embs = []

    test_transform = trans.Compose([
        trans.ToTensor(),
        trans.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5])])
    if bboxes.size != 0:
        for img in faces:
            if tta:
                mirror = cv2.flip(img, 1)
                emb = detect_model(test_transform(img).to(device).unsqueeze(0))
                emb_mirror = detect_model(test_transform(mirror).to(device).unsqueeze(0))
                embs.append(l2_norm(emb + emb_mirror))
            else:
                embs.append(detect_model(test_transform(img).to(device).unsqueeze(0)))
        source_embs = torch.cat(embs)  # number of detected faces x 512

        diff = source_embs.unsqueeze(-1) - targets.transpose(1, 0).unsqueeze(0)  # i.e. 3 x 512 x 1 - 1 x 512 x 2 = 3 x 512 x 2
        dist = torch.sum(torch.pow(diff, 2), dim=1)  # number of detected faces x numer of target faces
        # print("dist", dist)
        minimum, min_idx = torch.min(dist, dim=1)  # min and idx for each row

        min_idx[minimum > ((threshold - 156) / (-80))] = -1  # if no match, set idx to -1
        # cos_dis = calculate_cosine_distance(source_embs,targets)
        # print(cos_dis)
        # minimum, min_idx = torch.min(cos_dis, dim=1) # min and idx for each row
        # print("min_idx",min_idx)
        # min_idx[minimum > ((threshold-156)/(-80))] = -1  # if no match, set idx to -1
        score = minimum
        results = min_idx
        for id in min_idx:
            if id != -1:
                res.append(mssv[id])
        # print(score / torch.mean(dist, dim=1))
        # convert distance to score dis(0.7,1.2) to score(100,60)
        score_100 = torch.clamp(score * -80 + 156, 0, 100)
    image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype('Weights/times.ttf', 30)

    FPS = 1.0 / (time.time() - start_time)
    draw.text((10, 10), 'FPS: {:.1f}'.format(FPS), fill=(0, 0, 0), font=font)
    if bboxes.size != 0 :
        for i, b in enumerate(bboxes):
            draw.rectangle([(b[0], b[1]), (b[2], b[3])], outline='blue', width=5)
            if score_label:
                if score_100[i]>70:
                    draw.text((int(b[0]), int(b[1] - 25)), names[results[i]] + ' score:{:.0f}'.format(score_100[i]),
                          fill=(255, 255, 0), font=font)
                    print(names[results[i]])
                    print(dist)
                else:
                    # results[i]="unknow"
                    draw.text((int(b[0]), int(b[1] - 25)), "unknow", fill=(255, 255, 0), font=font)
            else:
                draw.text((int(b[0]), int(b[1] - 25)), names[results[i]], fill=(255, 255, 0), font=font)
                print(names[results[i]])

        for p in landmarks:
            for i in range(5):
                draw.ellipse([(p[i] - 2.0, p[i + 5] - 2.0), (p[i] + 2.0, p[i + 5] + 2.0)], outline='blue')
    if fake_face.size !=0:
        for i, b in enumerate(fake_face):
            draw.rectangle([(b[0], b[1]), (b[2], b[3])], outline='red', width=5)
            draw.text((int(b[0]), int(b[1] - 25)), "fake face", fill=(255, 255, 0), font=font)
        if bboxes.size==0:
            frame = cv2.cvtColor(np.asarray(image), cv2.COLOR_RGB2BGR)
            return frame,None,None
    frame = cv2.cvtColor(np.asarray(image), cv2.COLOR_RGB2BGR)

    return frame, res,score_100



if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='face detection demo')
    parser.add_argument('-th','--threshold',help='threshold score to decide identical faces',default=60, type=float)
    parser.add_argument("-u", "--update", help="whether perform update the facebank",action="store_true", default= False)
    parser.add_argument("-tta", "--tta", help="whether test time augmentation",action="store_true", default= True)
    parser.add_argument("-c", "--score", help="whether show the confidence score",action="store_true",default= True )
    parser.add_argument("--scale", dest='scale', help="input frame scale to accurate the speed", default=0.5, type=float)
    parser.add_argument('--mini_face', dest='mini_face', help="Minimum face to be detected. derease to increase accuracy. Increase to increase speed",
                        default=20, type=int)
    args = parser.parse_args()
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

    detect_model = MobileFaceNet(512).to(device)  # embeding size is 512 (feature vector)
    detect_model.load_state_dict(torch.load('Weights/MobileFace_Net', map_location=lambda storage, loc: storage))
    print('MobileFaceNet face detection model generated')
    detect_model.eval()

    if args.update:
        targets, names = prepare_facebank(detect_model, path='D21CQCN01', tta=args.tta)
        print('facebank updated')
    else:
        targets, names = load_facebank(path='D21CQCN01')
        print('facebank loaded')
        print(names[36])
        print(targets.shape)
        # targets: number of candidate x 512

    cap = cv2.VideoCapture(0)
    cap.set(cv2.CAP_PROP_AUTO_EXPOSURE, 1)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
    while True:
        isSuccess, frame = cap.read()
        if isSuccess:
            try:
                start_time = time.time()
                input = resize_image(frame, args.scale)
                bboxes, landmarks = create_mtcnn_net(input, args.mini_face, device, p_model_path='Weights/pnet_Weights',
                                                     r_model_path='Weights/rnet_Weights',
                                                     o_model_path='Weights/onet_Weights')

                if bboxes != []:
                    bboxes = bboxes / args.scale
                    landmarks = landmarks / args.scale

                faces = Face_alignment(frame, default_square=True, landmarks=landmarks)

                embs = []

                test_transform = trans.Compose([
                                trans.ToTensor(),
                                trans.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5])])

                for img in faces:

                    if args.tta:

                        mirror = cv2.flip(img,1)

                        emb = detect_model(test_transform(img).to(device).unsqueeze(0))
                        emb_mirror = detect_model(test_transform(mirror).to(device).unsqueeze(0))
                        embs.append(l2_norm(emb + emb_mirror))
                    else:
                        embs.append(detect_model(test_transform(img).to(device).unsqueeze(0)))

                source_embs = torch.cat(embs)  # number of detected faces x 512
                print(source_embs.shape)
                print(targets.shape)
                diff = source_embs.unsqueeze(-1) - targets.transpose(1, 0).unsqueeze(0) # i.e. 3 x 512 x 1 - 1 x 512 x 2 = 3 x 512 x 2
                dist = torch.sum(torch.pow(diff, 2), dim=1) # number of detected faces x numer of target faces
                print("dist",dist)

                minimum, min_idx = torch.min(dist, dim=1) # min and idx for each row
                print("min_idx",min_idx)
                min_idx[minimum > ((args.threshold-156)/(-80))] = -1  # if no match, set idx to -1
                # cos_dis = calculate_cosine_distance(source_embs,targets)
                # print(cos_dis)
                # minimum, min_idx = torch.min(cos_dis, dim=1) # min and idx for each row
                # print("min_idx",min_idx)
                # min_idx[minimum > ((args.threshold-156)/(-80))] = -1  # if no match, set idx to -1
                score = minimum
                results = min_idx
                print(score/torch.mean(dist, dim=1))
                # convert distance to score dis(0.7,1.2) to score(100,60)
                score_100 = torch.clamp(score*-80+156,0,100)

                image = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                draw = ImageDraw.Draw(image)
                font = ImageFont.truetype('Weights/simkai.ttf', 30)

                FPS = 1.0 / (time.time() - start_time)
                draw.text((10, 10), 'FPS: {:.1f}'.format(FPS), fill=(0, 0, 0), font=font)

                for i, b in enumerate(bboxes):
                    draw.rectangle([(b[0], b[1]), (b[2], b[3])], outline='blue', width=5)
                    if args.score:

                        draw.text((int(b[0]), int(b[1]-25)), names[results[i] ] + ' score:{:.0f}'.format(score_100[i]), fill=(255,255,0), font=font)
                    else:
                        draw.text((int(b[0]), int(b[1]-25)), names[results[i] ], fill=(255,255,0), font=font)
                        print(names[results[i]])

                for p in landmarks:
                    for i in range(5):
                        draw.ellipse([(p[i] - 2.0, p[i + 5] - 2.0), (p[i] + 2.0, p[i + 5] + 2.0)], outline='blue')

                frame = cv2.cvtColor(np.asarray(image), cv2.COLOR_RGB2BGR)

            except:
                print('detect error')

            cv2.imshow('video', frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()