#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 21 09:09:25 2019
Generate the face bank

@author: AIRocker
"""

import sys
import os
# sys.path.append(os.path.join(sys.path[0], 'MTCNN'))
from MTCNN import create_mtcnn_net,create_mtcnn_net_facebank
from align_trans import *
import numpy as np
from torchvision import transforms as trans
import torch
from face_model import MobileFaceNet, l2_norm
from pathlib import Path
import cv2
import pickle
test_transform = trans.Compose([
    trans.ToTensor(),
    trans.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5])])

device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
from openpyxl import Workbook,load_workbook
import os
from openpyxl.utils import get_column_letter
import datetime
day = datetime.date.today()

def create_excel(lop, dict_class):
    current_path = os.path.abspath(__file__)
    current_path = os.path.dirname(current_path)
    current_path = os.path.join(current_path,lop)
    filename = os.path.join(current_path,"list.xlsx")
    sorted_list = sorted([[key, value["name"]] for key, value in dict_class.items()], key=lambda x: x[1])
    if  os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        # Ghi dữ liệu vào file đã có
    else:
        # Ghi dữ liệu vào file mới
        wb = Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="MSSV")
        ws.cell(row =2,column =2, value="Họ và tên")
        for i in range(len(sorted_list)):
            ws['A'+str(i+3)]=sorted_list[i][0]
            ws['B' + str(i + 3)] = sorted_list[i][1]
        for column in ws.columns:
            column_letter = get_column_letter(column[0].column)
            lengths = [len(str(cell.value)) for cell in column]
            max_length = max(lengths)
            adjusted_width = (max_length + 2) * 1.2  # Thêm khoảng trắng và margin
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(filename)
# create_excel(imgs_path,ds_mssv,ds_ten)
def prepare_facebank(path , tta = True):
    model = MobileFaceNet(512).to(device)  # embeding size is 512 (feature vector)
    model.load_state_dict(
        torch.load('Weights/MobileFace_Net', map_location=lambda storage, loc: storage))
    print('MobileFaceNet face detection model generated')
    model.eval()
    info = {}
    list_dir = os.listdir(path)
    print(list_dir)
    embs = []
    error_image =[]
    for files in list_dir:
        image_path = os.path.join(path, files)
        try :
            img = cv2.imdecode(np.fromfile(image_path, dtype=np.uint8), cv2.IMREAD_UNCHANGED)

            if img.shape != None:
                bboxes, landmarks = create_mtcnn_net_facebank(img, 20, device,
                                                 p_model_path='Weights/pnet_Weights',
                                                 r_model_path='Weights/rnet_Weights',
                                                 o_model_path='Weights/onet_Weights')

            img = Face_alignment(img, default_square=True, landmarks=landmarks)

            with torch.no_grad():
                if tta:
                    mirror = cv2.flip(img[0], 1)
                    emb = model(test_transform(img[0]).to(device).unsqueeze(0))
                    emb_mirror = model(test_transform(mirror).to(device).unsqueeze(0))
                    res =l2_norm(emb + emb_mirror)
                else:
                    res =model(test_transform(img[0]).to(device).unsqueeze(0))
            name=files.split("-")[1].split(".")[0]
            mssv = files.split("-")[0]
            # print("res",res)
            info[mssv]= {"name":name,"emb":res}
        except Exception as e:
            print(e)
            error_image.append(image_path)
    #giảm chiều


    # lấy đường dẫn tới file hiện tại
    current_path = os.path.abspath(__file__)
    current_path = os.path.dirname(current_path)

    # lấy tên của  lớp
    folder_name = os.path.basename(path)
    # save VD :D21CQCN01_emb_mobile.pth file  nhúng
    current_path = os.path.join(current_path,  "Config")

    if not os.path.exists(current_path):
        os.mkdir(current_path)
    save_facebank(info,folder_name)
    save_class(folder_name)
    return error_image
def save_class(class_name):
    if not os.path.exists("class.txt"):
        with open("class.txt","w") as file:
            file.write(class_name + "\n")
    else:
        with open("class.txt", "a") as file:
            file.write(class_name + "\n")

def save_facebank(data,name_class):
    current_path = os.path.abspath(__file__)
    current_path = os.path.dirname(current_path)
    current_path = os.path.join(current_path,"Config")
    file_path =os.path.join(current_path, name_class+".pkl")
    with open(file_path, "wb") as file:
        pickle.dump(data, file)

def load_facebank(name_class):
    data_path = Path("Config")
    data_path = os.path.join(data_path,name_class+".pkl")
    with open(data_path, "rb") as file:
        loaded_dict = pickle.load(file)
    names = []
    keys = []
    values = []

    for key, value in loaded_dict.items():
        names.append(value['name'])
        keys.append(key)
        values.append(value['emb'])

    # Stack các tensor values
    stacked_values = torch.stack(values).squeeze()

    # Sắp xếp theo name
    sorted_indices = sorted(range(len(names)), key=lambda x: names[x])
    names = [names[i] for i in sorted_indices]
    keys = [keys[i] for i in sorted_indices]
    stacked_values = stacked_values[sorted_indices]
    return stacked_values,names,keys

def info_class(name_class):
    data_path = Path("Config")
    data_path = os.path.join(data_path,name_class+".pkl")
    with open(data_path, "rb") as file:
        loaded_dict = pickle.load(file)
    return loaded_dict
def update_class(name_class,link_folder):
    loaded_dict = info_class(name_class)
    name_header = list(loaded_dict[next(iter(loaded_dict))].keys())
    name_header.remove("emb")
    name_header.remove("name")
    model = MobileFaceNet(512).to(device)  # embeding size is 512 (feature vector)
    tta =True
    model.load_state_dict(
        torch.load('Weights/MobileFace_Net', map_location=lambda storage, loc: storage))
    print('MobileFaceNet face detection model generated')
    model.eval()
    list_dir = os.listdir(link_folder)
    print(list_dir)
    error_image = []
    for files in list_dir:
        image_path = os.path.join(link_folder, files)
        print("image_path",image_path)
        try:
            img = cv2.imdecode(np.fromfile(image_path, dtype=np.uint8), cv2.IMREAD_UNCHANGED)
            print(img)
            print("imgsshpe",img.shape)
            if img.shape != None:
                bboxes, landmarks = create_mtcnn_net_facebank(img, 20, device,
                                                     p_model_path='Weights/pnet_Weights',
                                                     r_model_path='Weights/rnet_Weights',
                                                     o_model_path='Weights/onet_Weights')

            img = Face_alignment(img, default_square=True, landmarks=landmarks)

            with torch.no_grad():
                if tta:
                    mirror = cv2.flip(img[0], 1)
                    emb = model(test_transform(img[0]).to(device).unsqueeze(0))
                    emb_mirror = model(test_transform(mirror).to(device).unsqueeze(0))
                    res =l2_norm(emb + emb_mirror)
                else:
                    res =model(test_transform(img[0]).to(device).unsqueeze(0))
            name = files.split("-")[1].split(".")[0]
            mssv = files.split("-")[0]
            loaded_dict[mssv] = {"name": name, "emb": res}
            for date in name_header:
                loaded_dict[mssv][date] = 0
            save_facebank(loaded_dict,name_class)

        except Exception as e:
            print(e)
            error_image.append(image_path)
    return error_image
def save_attendance(day, list_diemdanh, list_mssv, lop):
    dict_class = info_class(lop)
    for mssv  in list_mssv:
        dict_class[mssv][day]= 0
    for mssv in list_diemdanh:
        dict_class[mssv][day]=1
    save_facebank(dict_class,lop)

if __name__ == '__main__':
    a,b,c=load_facebank("D21CQCN01")
    print(a)
    # prepare_facebank("D21CQCN01")